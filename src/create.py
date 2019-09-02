import openpyxl
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfbase.ttfonts import TTFont

pdfmetrics.registerFont(TTFont('Arial','Arial.ttf'))
wb = openpyxl.load_workbook('C:\\Users\\vicky\\PycharmProjects\\payslip\\src\\PaySlip.xlsx')
sheet = wb['employees']

# print(sheet.cell(2, 2).value)
# print(sheet.cell(2, 3).value)

page_width = 1080
page_height = 1920
spread = 100
start = 200
start_2 = 600
company_name = 'Perficient'
month_year = 'September 2019'


def create_payslip():
    for i in range(2, 5):
        emp_id = sheet.cell(row=i, column=1).value
        emp_name = sheet.cell(row=i, column=2).value
        emp_lastname = sheet.cell(row=i, column=3).value
        gross_salary = sheet.cell(row=i, column=4).value
        bonus_payment = sheet.cell(row=i, column=8).value
        net_salary = sheet.cell(row=i, column=10).value

        c = canvas.Canvas(str(emp_name) + '_' + str(emp_lastname) + '_' + str(emp_id) + '_' + month_year + '.pdf')
        c.setPageSize((page_width, page_height))
        c.setFont('Arial', 70)
        text_width = stringWidth(company_name, 'Arial', 70)
        c.drawString((page_width - text_width)/2, 1800, company_name)

        text = 'Salary Calculation for period ' + month_year
        text_width = stringWidth(text, 'Arial', 35)
        c.setFont('Arial', 35)
        c.drawString((page_width - text_width) / 2, 1600, text)

        y = 1500

        c.setFont('Arial', 25)
        c.drawString(start, y, 'Employee\'s ID: ')
        c.drawString(start_2, y, str(emp_id))
        y -= spread
        c.drawString(start, y, 'Employee\'s Name: ')
        c.drawString(start_2, y, str(emp_name))
        y -= spread
        c.drawString(start, y, 'Employee\'s LastName: ')
        c.drawString(start_2, y, str(emp_lastname))
        y -= spread
        c.drawString(start, y, 'Employee\'s GrossSalary: ')
        c.drawString(start_2, y, str(gross_salary))
        y -= spread
        c.drawString(start, y, 'Employee\'s BonusAmount: ')
        c.drawString(start_2, y, str(bonus_payment))
        y -= spread
        c.drawString(start, y, 'Employee\'s NetSalary: ')
        c.drawString(start_2, y, str(net_salary))
        y -= spread * 3
        c.drawString(start, y, 'Employee\'s Signature: ')
        c.drawString(start_2, y, '________________')
        c.save()


create_payslip()
